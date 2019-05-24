#!/usr/bin/env python3
# -*- coding: UTF-8 -*-.

from openpyxl import load_workbook as open_xlsx
from openpyxl.cell import Cell as XlsxCell
from cgi import escape
from urllib.request import urlopen
from shutil import copyfileobj

import numpy as np
import os
import re
import datetime

from MPG.utils import get_sun, get_period_limits, iter_period_dates, load_config
from MPG.esolog import BasicLog
from MPG.esoarchive import NightRequest
from MPG.programlist import ProgramList

def get_shift_name(tel, period, path='.'):
    name = 'P{}-shift.htm'.format(period)
    path = BasicLog(tel=tel, period=period, path=path).get_path()
    return os.path.join(path, name)

def get_schedule_name(tel, period, format='xls', path='.'):
    if format in ['xlsx', 'xls']:
        name =  'P{}.{}'.format(period, format)
    elif format == 'html':
        name = 'P{}-schedule.htm'.format(period)
    else:
        raise RuntimeError('unknown format: ' + str(format))
    path = BasicLog(tel=tel, period=period, path=path).get_path()
    return os.path.join(path, name)


def get_schedule_url(tel, period, path='.'):
    config = load_config(tel, period, path=path)['Origin']
    hostname, urlpath = config['hostname'], config['path']
    if 'format' in config:
        fmt = config['format']
    else:
        fmt = 'xlsx'
    url = 'http://{}/{}/P{}.{}'.format(hostname, urlpath, period, fmt)
    return url


def load_xlsx_schedule(tel, period, path='.'):
    config = load_config(tel, period, path=path)
    origin = config['Origin']
    fmt = origin['format']
    if fmt != 'xlsx':
        raise RuntimeError('Schedule format is not xlsx')
    filename = get_schedule_name(tel, period, format=fmt, path=path)
    book = open_xlsx(filename)    
    if 'sheetname' in origin:
        sheetname = origin['sheetname']
    else:
        sheetname = 'P{0}'.format(period)
    sheet = book[sheetname]
    return sheet

class prog_sort:
    def __init__(self, mon):
        self.mon = mon
    def __call__(self, p):
        if p in re.split(', *', self.mon['daily']):
            return 2
        if p == 'Calib':
            return 0
        return 1

def merge_cells(row, config, night_length=8):
    # add specified daily monitoring programmes
    mon = config['Monitoring'] 
    if not any([e in row for e in re.split(', *', mon['excludes'])]):
        for p in re.split(', *', mon['daily']):
            if p not in row:
                row = row + [p]
    # unique 
    u, ind = np.unique(row, return_index=True)
    row = u[np.argsort(ind)].tolist()
    # sort programmes
    key = prog_sort(mon)
    row.sort(key=key)
    # give lengths in hours to programmes
    row = [[p, mon.getint(p, 1)] for p in row] 
    # complement unil length is min_length
    i = 0
    while sum(p[1] for p in row) < night_length:
        while row[i][0] in mon:
            i = (i + 1) % len(row)
        row[i][1] += 1
        i = (i + 1) % len(row)
    # split some coupled programmes
    i = 0
    while i < len(row):
        if ',' not in row[i][0]:
            i += 1
            continue
        progs = re.split(', *', row[i][0])
        nprogs = len(progs)
        totlen = row[i][1]
        indlen = totlen // nprogs
        row[i] = [progs[0], totlen - (nprogs - 1) * indlen]
        for j, p in enumerate(progs[1:]):
            row.insert(i + 1 + j, [p, indlen])
        i += nprogs 
    return row

def parse_xlsx_schedule(tel, period, path='.'):
    sheet = load_xlsx_schedule(tel, period, path=path)
    # configuration
    config = load_config(tel, period, path=path)    
    bg = config['BackgroundColour']
    spelling = config['Spelling']
    pos = config['Positions']
    rows = pos['rows']
    obs_cols = pos['observer_cols']
    sa_cols = pos['support_cols']
    prog_cols = pos['prog_cols']
    header_row = pos['header_row']
    # observer and service astronomer
    obs = get_values(sheet, rows, obs_cols, spelling=spelling)
    obs = [', '.join(row) for row in obs]
    sa = get_values(sheet, rows, sa_cols, spelling=spelling)
    sa = [', '.join(row) for row in sa]
    # typical night lengths (Winter/Summer) 
    night_length = 8 + 3 * (period % 2)
    # programmes
    date_start, date_end = get_period_limits(period)
    progs = get_values(sheet, rows, prog_cols, 
            header_row=header_row, header_col=0, bg=bg, spelling=spelling)
    progs = [merge_cells(row, config, night_length=night_length) 
                    for row in progs] 
    dates = [date for date in iter_period_dates(period)]
    ephem = [get_sun(date) for date in dates]
    return dates, ephem, obs, sa, progs

def write_html_shifts(data, tel, period, path='.'):
    dates, ephem, obs, sa, progs = data
    #for n, s in zip(dates, sa):
    #    print(n.strftime("%Y-%m-%d"), re.split(', *', s))
    config = load_config(tel, period, path=path)
    positions = config['Positions']
    supports = positions.get('supports').split(',')
    day = datetime.timedelta(days=1)
    html = '<table class="horizontal">\n'
    url = get_schedule_url(tel, period, path=path)
    today = datetime.datetime.today().strftime('%Y-%b-%d')
    begin = dates[0].strftime('%b %Y')
    end = dates[-1].strftime('%b %Y')
    html += '    <caption>Schedule for ESO period {period} (covering {begin}-{end}). It has been automatically generated from the <a href="{url}">official schedule file</a> on {today}.</caption>\n'.format(period=period, begin=begin, end=end, url=url, today=today)
    html += '    <thead>\n'
    html += '        <tr>\n'
    html += get_cell_html('support astronomer', type_='th')
    html += get_cell_html('number of nights', type_='th')
    html += get_cell_html('shift start (noon)', type_='th')
    html += get_cell_html('shift end (noon)', type_='th')
    html += '        </tr>\n'
    html += '    </thead>\n'
    for support in supports:
        html += '    <tbody>\n'
        nights = [n for n, s in zip(dates, sa) 
            if support in re.split(', *', s)] 
        if len(nights) == 0:
            print('WARNING: ' + support + ' has no shifts')
            continue
        index = np.where(np.diff(nights) != day)[0]
        beg_index = np.hstack([0, index + 1])
        end_index = np.hstack([index, len(nights) - 1])
        nshift = len(beg_index)
        if nshift == 0:
            continue
        nnight = len(nights)
        print('{} has {} nights in {} shifts'.format(support, nnight, nshift))
        for  i, (b, e) in enumerate(zip(beg_index, end_index)):
            html += '        <tr>\n'
            if i == 0:
                html += get_cell_html(support, rowspan=nshift)
                html += get_cell_html(nnight, rowspan=nshift)
            begin = nights[b].strftime('%Y-%m-%d')
            end = (nights[e] + day).strftime('%Y-%m-%d')
            html += get_cell_html(begin)
            html += get_cell_html(end)
            html += '        </tr>\n'
        html += '    </tbody>\n'
    html += '</table>\n'
    filename = get_shift_name(tel, period, path=path)
    with open(filename, 'w') as fh:
        fh.write(html)

def write_html_schedule(data, tel, period, path='.'):
    dates, ephem, obs, sa, progs = data
    night_length = 8 + 3 * (period % 2)
    # write
    programs = ProgramList(tel, period, path=path, honour_omit=False)
    url = get_schedule_url(tel, period, path=path)
    today = datetime.datetime.today().strftime('%Y-%b-%d')
    begin = dates[0].strftime('%b %Y')
    end = dates[-1].strftime('%b %Y')
    html = '<table class="horizontal">\n'
    html += '    <caption>Schedule for ESO period {period} (covering {begin}-{end}). It has been automatically generated from the <a href="{url}">official schedule file</a> on {today}.</caption>\n'.format(period=period, begin=begin, end=end, url=url, today=today)
    html += '    <thead>\n'
    html += '        <tr>\n'
    html += get_cell_html('date', type_='th') 
    html += get_cell_html(('sun', 2), type_='th')  
    html += get_cell_html(('night', 2), type_='th') 
    html += get_cell_html('support', type_='th') 
    html += get_cell_html('visitor', type_='th') 
    html += get_cell_html(('programmes', night_length), type_='th') 
    html += '        </tr>\n'
    html += '        <tr class="suptitle">\n'
    html += get_cell_html('[data]', type_='th') 
    html += get_cell_html('begin', type_='th') 
    html += get_cell_html('end', type_='th') 
    html += get_cell_html('begin', type_='th') 
    html += get_cell_html('end', type_='th') 
    for i in range(night_length + 2):
        html += get_cell_html('', type_='th') 
    html += '        </tr>\n'
    html += '    </thead>\n'
    html += '    <tbody>\n'
    for d, e, s, o, p in zip(dates, ephem, sa, obs, progs):
        html += '        <tr class="weekday{}">\n'.format(d.weekday())
        html += get_date_html(d) 
        for t in [e[0], e[3], e[1], e[2]]:
            html += get_ephem_html(t) 
        html += get_cell_html(s)
        html += get_cell_html(o) 
        for pi in p:
            html += get_prog_html(pi, programs)
        html += '        </tr>\n'
    html += '    </tbody>\n'
    html += '</table>\n'
    filename = get_schedule_name(tel, period, format='html', path=path)
    with open(filename, 'w') as fh:
        fh.write(html)

def spell_check(value, spelling):
    if value in spelling:
        return spelling[value]
    return value

def get_values(sheet, rows, cols, header_row=None, header_col=None, 
            remove_empty=True, **keyw):
    r1, r2 = rows.split(':')
    if ':' in cols:
        c1, c2 = cols.split(':')
    else:
        c1, c2 = cols, cols
    rng = c1 + r1 + ':' + c2 + r2
    area = sheet[rng]
    # Column and row header
    colhead = [None] * len(area[0]) 
    if header_row is not None:
        header_rng = c1 + str(header_row) + ':' + c2 + str(header_row)
        colhead = sheet[header_rng][0]
    rowhead = [None] * len(area)
    if header_col is not None:
        rowhead = list(zip(*area))[header_col]
    # Get individual values, possibly replacing by header value if empty
    # and same background colour.
    values = [[get_cell_value(cell, headers=[ch, rh], **keyw) 
                    for cell, ch in zip(row, colhead)] 
                            for row, rh in zip(area, rowhead)]
    if remove_empty:
        values = [[r for r in row if r != ''] for row in values]
    return values 
 
def get_cell_value(cell, headers=[], bg=None, spelling=None, verbose=False):
    val = cell.value
    if val is None:
        val = '' 
    val = str(val).strip()
    val = re.sub('^\\[', '', val)
    val = re.sub('\\]$', '', val)
    cell_bg = cell.fill.fgColor
    if verbose and val != '':
        print('        not empty cell', val)
    # if an empty cell has the same RGB colour as a given programme
    # replace it
    if bg is not None and cell_bg.type == 'rgb':
        if cell_bg.rgb in bg:
            val = bg[cell_bg.rgb]
            if verbose:
                print('        -> retrieve config value using bg: ', val)
    # if an empty cell has the same colour as the column/row header
    # it should have the same contents
    for header in headers:
        if header is not None and val == '':  
            header_bg = header.fill.fgColor
            header_val = get_cell_value(header, verbose=False)
            if val == '' and xlsx_same_color(cell_bg, header_bg):
                val = header_val
                if verbose:
                    print('        -> retrieve header value: ', val)
    if spelling is not None:
        if val in spelling:
            val = spelling[val]
    return val
   
def xlsx_same_color(a, b):
    if isinstance(a, XlsxCell):
        a = a.fill.fgColor
    if isinstance(b, XlsxCell):
        b = b.fill.fgColor
    # no colour set
    if ((a.type == 'theme' and a.theme == 0 and a.tint == 0 or 
            a.type == 'rgb' and a.rgb == '00000000') 
       and (b.type == 'theme' and b.theme == 0 and a.tint == 0 or
            b.type == 'rgb' and b.rgb == '00000000')):
        return True
    # if a colour is set
    if a.type != b.type:
        return False
    if a.type == 'theme':
        return a.theme == b.theme and abs(a.tint - b.tint) < 1e-6
    if a.type == 'rgb':
        return a.rgb == b.rgb
    return False 
 
def get_date_html(date, indent=3, cr='\n'):
    d1, d2 = date, date + datetime.timedelta(days=1)
    night = d1.isoformat()
    d1, d2 = [d.strftime('%A %B %d') for d in [d1, d2]]
    link = NightRequest(night, inslist=['FEROS', 'WFI', 'GROND']).url()
    html = '<td title="night from {} to {}"><a href="{}">{}</a></td>'.format(d1, d2, link, night)
    html = ('    ' * indent) + html + cr
    return html

def get_ephem_html(e, indent=3, cr='\n'):
    h, m, s = e.timetuple()[3:6]
    if s > 30:
        m += 1
        if m == 60:
            m = 0
            h += 1
            if h == 24:
                h = 0
    html = '<td>{:02}:{:02}</td>'.format(h, m)
    html = ('    ' * indent) + html + cr
    return html

def get_cell_html(x, type_='td', rowspan=1, indent=3, cr='\n'):
    if not isinstance(x, tuple):
        x = (x, 1)
    value, colspan = x
    span = ''
    if colspan > 1:
        span += 'colspan="{}" '.format(colspan)
    if rowspan > 1:
        span += 'rowspan="{}" '.format(rowspan)
    html = '<{} {} class="atomic">{}</td>'.format(type_, span, value)
    html = ('    ' * indent) + html + cr
    return html

def get_prog_html(x, programs, indent=3, cr='\n'):
    tc = ''
    colspan = ''
    p, n = x
    if n > 1:
        colspan = 'colspan="{}" '.format(n)
    if p == '':
        return '<td {}></td>'.format(ctype, colspan)
    prog = programs.lookup(p)
    cls = prog['TAC'].lower()
    print(cls[0:6])
    if cls[0:6] == 'mpia/p':
        cls = 'mpiacomp'
    print(cls)
    cls = re.sub('/', '', cls)
    if prog['Time-critical'] == 'yes':
        cls += ' timecritical'
        tc = ' (time-critical)'
    if 'ToO' in prog['PID']:
        cls += ' buffer'
    title = escape('{}{} - {}'.format(prog['TAC'], tc, prog['Title']))
    html = '<td {}class="{}" title="{}">{}</td>'.format(
                colspan, cls, title, p)
    html = ('    ' * indent) + html + cr
    return html

def publish_html_schedule(tel, period, rdir='lachaume@black:~/2.2m', path='.'):
    sched = get_schedule_name(tel, period, format='html', path=path)
    shift = get_shift_name(tel, period, path=path)
    cmd = 'scp {} {}/schedule/P{}'.format(sched, rdir, period)
    os.system(cmd)
    cmd = 'scp {} {}/support/P{}'.format(shift, rdir, period)
    os.system(cmd)

def retrieve_xls_schedule(tel, period, path='.'):
    config = load_config(tel, period, path=path)['Origin']
    hostname, urlpath = config['hostname'], config['path']
    if 'format' in config:
        fmt = config['format']
    else:
        fmt = 'xls'
    if path.lower() != 'none':
        locname = get_schedule_name(tel, period, format=fmt, path=path)
        url = 'http://{}/{}'.format(hostname, urlpath, period, fmt)
        with urlopen(url) as response, open(locname, 'wb') as out:
            copyfileobj(response, out)
        if fmt != 'xls':
           xls = get_schedule_name(tel, period, format='xls', path=path)
           cmd = '/usr/bin/ssconvert -T Gnumeric_Excel:excel_dsf'
           os.system('{} {} {} 2>/dev/null'.format(cmd, locname, xls))


if __name__ == "__main__":
    tel = '2.2m'
    period = 103
    path = "/home/regis/Work/2.2m/nightlogs"
    config = load_config(tel=tel, period=period, path=path)
    row = ['Jenkins, Bayliss', 'Calib']
    #retrieve_xls_schedule(tel, period, path=path)
    #sheet = load_xlsx_schedule(tel, period, path=path) 
    #data = parse_xlsx_schedule(tel, period, path=path)
    #write_html_schedule(data, tel, period, path=path)
    #write_html_shifts(data, tel, period, path=path)
    #publish_html_schedule(tel, period)
